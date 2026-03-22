from __future__ import annotations

import logging
import math
import os
import time
from dataclasses import dataclass
from datetime import datetime
from typing import List

from core.thread_management.registry import registry

logger = logging.getLogger(__name__)

TARGET_DEFAULT: float = -3.0
STOP_THRESHOLD: float = 0.05
KP: float = 0.07
KD: float = 0.015
BRAKE_FF_GAIN: float = 0.10         # brake pedal fraction per m/s² at reference weight
REFERENCE_WEIGHT_KG: float = 8500.0 # unloaded truck baseline for FF scaling
MIN_BRAKE: float = 0.02
GRAVITY: float = 9.81
RESULTS_FILE: str = os.path.join("logs", "decel_test_results.xlsx")


def _idle_cancel(speed_ms: float) -> float:
    s = abs(speed_ms)
    return max(
        ((1 - ((s + 0.1) / 5.5) ** 2.22) * (-0.6 / (s + 0.2) + 1)) / 0.715 * 0.10,
        0.0,
    )


def _calc_truck_weight(cargo_mass: float, has_trailer: bool) -> float:
    """
    Estimate total vehicle mass in kg.

    unitMass (trailer body) and fuel are not exposed by the current telemetry
    thread, so they are approximated with constants.  Update when those fields
    become available.
    """
    unit_mass = 8500.0  # truck tractor + approximate fuel mass
    if has_trailer:
        return unit_mass + cargo_mass
    return unit_mass


@dataclass
class DecelSample:
    t:              float
    speed:          float
    brake_out:      float
    measured_decel: float
    target_decel:   float
    rotationY:      float
    gravity_effect: float   # m/s², positive = uphill (reduces brake demand)
    idle_offset:    float
    weight_kg:      float


class DecelController:
    """
    Closed-loop deceleration controller.

    Owned by SendingThread and ticked synchronously each loop cycle so brake
    output is applied without queue delay.

    tick(target_ms2) reads all required state from the telemetry thread via the
    registry, so callers (ACC, test harness) only need to supply the desired
    deceleration rate.

    Usage from SendingThread:
        brake = self._decel.tick(target_ms2, self.loop_interval)
        if self._decel.active:
            controller.aforward  = 0.0
            controller.abackward = brake

    Usage from ACC thread:
        sending = registry.get_thread("sending_thread")
        sending.start_decel(-2.5)
        sending.stop_decel()
    """

    def __init__(self) -> None:
        self._active: bool = False
        self._start_time: float = 0.0
        self._last_time: float = 0.0
        self._last_speed: float = 0.0
        self._last_error: float = 0.0
        self._samples: List[DecelSample] = []

    @property
    def active(self) -> bool:
        return self._active

    @property
    def last_sample(self) -> DecelSample | None:
        return self._samples[-1] if self._samples else None

    def start(self, speed: float) -> None:
        self._active = True
        self._start_time = time.monotonic()
        self._last_time = 0.0
        self._last_speed = speed
        self._last_error = 0.0
        self._samples = []
        logger.info("decel controller started from %.2f m/s", speed, extra={"popup": True})

    def cancel(self) -> None:
        if self._active:
            logger.info("decel controller cancelled")
        self._active = False
        self._samples = []

    def tick(self, target_ms2: float, loop_interval: float) -> float:
        """
        Read telemetry from registry, compute and return brake pedal [0..1].
        Automatically stops and saves results when speed <= STOP_THRESHOLD.
        Returns 0.0 when not active or telemetry is unavailable.
        """
        if not self._active:
            return 0.0

        speed, rotationY, cargo_mass, has_trailer, connected = self._read_telemetry()

        if not connected:
            logger.debug("decel controller: telemetry not connected, holding last brake")
            return self.last_sample.brake_out if self.last_sample else MIN_BRAKE

        now = time.monotonic()
        dt = (now - self._last_time) if self._last_time > 0.0 else loop_interval
        dt = max(dt, 1e-6)

        measured_decel = (speed - self._last_speed) / dt
        error = measured_decel - target_ms2
        d_error = (error - self._last_error) / dt

        # Gravity effect in m/s² — positive = uphill, reduces brake demand.
        # Mass cancels in F=ma so this is weight-independent.
        gravity_effect = GRAVITY * math.sin(rotationY)

        idle_offset = _idle_cancel(speed)

        weight_kg = _calc_truck_weight(cargo_mass, has_trailer)
        weight_factor = weight_kg / REFERENCE_WEIGHT_KG

        decel_from_brake = abs(target_ms2) - gravity_effect
        brake_ff = max(0.0, decel_from_brake) * BRAKE_FF_GAIN * weight_factor
        pd = KP * error + KD * d_error
        brake_out = min(1.0, max(MIN_BRAKE, brake_ff + pd + idle_offset))

        self._samples.append(DecelSample(
            t=now - self._start_time,
            speed=speed,
            brake_out=brake_out,
            measured_decel=measured_decel,
            target_decel=target_ms2,
            rotationY=rotationY,
            gravity_effect=gravity_effect,
            idle_offset=idle_offset,
            weight_kg=weight_kg,
        ))

        self._last_speed = speed
        self._last_time = now
        self._last_error = error

        if speed <= STOP_THRESHOLD:
            logger.info("decel controller: vehicle stopped — saving results")
            self._active = False
            self._save_excel(list(self._samples))
            self._samples = []

        return brake_out

    def _read_telemetry(self) -> tuple[float, float, float, bool, bool]:
        """Return (speed_ms, rotationY_rad, cargo_mass_kg, has_trailer, is_connected)."""
        try:
            tel = registry.get_thread("telemetry_thread")
            with tel.data._lock:
                return (
                    float(tel.data.speed),
                    float(tel.data.rotationY),
                    float(tel.data.cargoMass),
                    bool(tel.data.ego_has_trailer),
                    bool(tel.data.is_connected),
                )
        except (KeyError, AttributeError, TypeError) as exc:
            logger.debug("decel controller: telemetry read failed: %s", exc)
            return 0.0, 0.0, 0.0, False, False

    def _save_excel(self, samples: List[DecelSample]) -> None:
        if not samples:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.data_source import NumDataSource, NumRef
            from openpyxl.chart.series import Series, SeriesLabel
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error(
                "decel test: openpyxl not installed — cannot save results\npip install openpyxl",
                extra={"popup": True},
            )
            return

        try:
            os.makedirs(os.path.dirname(RESULTS_FILE) or ".", exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, _ = os.path.splitext(RESULTS_FILE)
            path = f"{base}_{ts}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Decel Test"

            headers = [
                "t (s)", "speed (m/s)", "brake_out",
                "measured_decel (m/s²)", "target_decel (m/s²)",
                "rotationY (rad)", "gravity_effect (m/s²)",
                "idle_offset", "decel_error (m/s²)", "weight_kg",
            ]
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for i, s in enumerate(samples, start=2):
                ws.cell(row=i, column=1,  value=round(s.t, 4))
                ws.cell(row=i, column=2,  value=round(s.speed, 4))
                ws.cell(row=i, column=3,  value=round(s.brake_out, 4))
                ws.cell(row=i, column=4,  value=round(s.measured_decel, 4))
                ws.cell(row=i, column=5,  value=round(s.target_decel, 4))
                ws.cell(row=i, column=6,  value=round(s.rotationY, 5))
                ws.cell(row=i, column=7,  value=round(s.gravity_effect, 4))
                ws.cell(row=i, column=8,  value=round(s.idle_offset, 4))
                ws.cell(row=i, column=9,  value=round(s.measured_decel - s.target_decel, 4))
                ws.cell(row=i, column=10, value=round(s.weight_kg, 1))

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

            avg_decel = sum(s.measured_decel for s in samples) / len(samples)
            max_brake = max(s.brake_out for s in samples)
            duration = samples[-1].t
            avg_weight = sum(s.weight_kg for s in samples) / len(samples)

            ws.cell(row=1, column=12, value="Summary").font = Font(bold=True)
            ws.cell(row=2, column=12, value="Duration (s)")
            ws.cell(row=2, column=13, value=round(duration, 3))
            ws.cell(row=3, column=12, value="Avg decel (m/s²)")
            ws.cell(row=3, column=13, value=round(avg_decel, 4))
            ws.cell(row=4, column=12, value="Target decel (m/s²)")
            ws.cell(row=4, column=13, value=samples[0].target_decel)
            ws.cell(row=5, column=12, value="Max brake pedal")
            ws.cell(row=5, column=13, value=round(max_brake, 4))
            ws.cell(row=6, column=12, value="Avg weight (kg)")
            ws.cell(row=6, column=13, value=round(avg_weight, 1))
            ws.cell(row=7, column=12, value="Samples")
            ws.cell(row=7, column=13, value=len(samples))

            chart = LineChart()
            chart.title = "Deceleration over time"
            chart.y_axis.title = "decel (m/s²)"
            chart.x_axis.title = "t (s)"
            chart.width = 20
            chart.height = 12

            def _make_series(col, label):
                f = f"'Decel Test'!${get_column_letter(col)}$2:${get_column_letter(col)}${len(samples) + 1}"
                s = Series()
                s.val = NumDataSource(numRef=NumRef(f=f))
                s.tx = SeriesLabel(v=label)
                return s

            chart.series.append(_make_series(4, "measured"))
            chart.series.append(_make_series(5, "target"))
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(samples) + 1))

            ws.add_chart(chart, "M2")
            wb.save(path)
            logger.info("decel test results saved → %s", path, extra={"popup": True})

        except Exception:
            logger.exception("decel test: failed to save Excel results")
